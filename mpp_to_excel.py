#!/usr/bin/env python3
"""
Convert a Microsoft Project .mpp file into the Project for the Web
"Excel Import Template" (Project / Resources / Tasks sheets), with the
Tasks data wrapped in an Excel table named "Estimates".

IMPORTANT — why this fills an embedded template instead of building a
workbook from scratch:
    The SQIT import plugin's Excel reader reads INLINE strings only. A
    workbook built fresh by openpyxl stores text in a shared-strings table
    (cells hold only integer indices), which the plugin reads as blank —
    the symptom is validation reporting "Project Name missing" and
    "Tasks: 0" even though the file clearly has data. Loading and filling
    the shipped template keeps every string cell as t="inlineStr", which
    the plugin reads correctly. Do not switch this back to openpyxl.Workbook().

Requirements:
    pip install mpxj openpyxl jpype1
"""

import sys
import glob
import os
import io
import base64
import datetime
import re

# ---------------------------------------------------------------------------
# Embedded empty import template (inline-string .xlsx). Filled at runtime.
# ---------------------------------------------------------------------------
_TEMPLATE_B64 = """
UEsDBBQAAAAIAM1xGF1Gx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZih6kDkQ9ip68zy51hbYpbYT67+0EP255ecgboi6J
Iia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5GZ3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+L
U8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIAM1xGF38s+k8+QAAACsCAAARAAAAZG9jUHJvcHMvY29yZS54bWzNksFKxDAQhl9Fcm+nydZ1Cd1eFE8KggXFW0hm
d4NNG5KRdt/etO52FX0AIZfM/PnmG0ilvdR9wKfQewxkMV6Nru2i1H7LDkReAkR9QKdinhJdau764BSla9iDV/pd7RFEUazBISmjSMEEzPxCZHVltNQBFfXh
hDd6wfuP0M4wowFbdNhRBJ5zYPU00R/HtoILYIIRBhe/CmgW4lz9Ezt3gJ2SY7RLahiGfFjNubQDh9fHh+d53cx2kVSnMb2KVtLR45adJ7+sbu+ae1aLQqyz
YpNx0fBrWW4kF2+T6w+/i7Drjd3Zf2EsyoaXMh1x8834LFhX8Otf1J9QSwMEFAAAAAgAzXEYXZlcnCMQBgAAnCcAABMAAAB4bC90aGVtZS90aGVtZTEueG1s
7Vpbc9o4FH7vr9B4Z/ZtC8Y2gba0E3Npdtu0mYTtTh+FEViNbHlkkYR/v0c2EMuWDe2STbqbPAQs6fvORUfn6Dh58+4uYuiGiJTyeGDZL9vWu7cv3uBXMiQR
QTAZp6/wwAqlTF61WmkAwzh9yRMSw9yCiwhLeBTL1lzgWxovI9bqtNvdVoRpbKEYR2RgfV4saEDQVFFab18gtOUfM/gVy1SNZaMBE1dBJrmItPL5bMX82t4+
Zc/pOh0ygW4wG1ggf85vp+ROWojhVMLEwGpnP1Zrx9HSSICCyX2UBbpJ9qPTFQgyDTs6nVjOdnz2xO2fjMradDRtGuDj8Xg4tsvSi3AcBOBRu57CnfRsv6RB
CbSjadBk2PbarpGmqo1TT9P3fd/rm2icCo1bT9Nrd93TjonGrdB4Db7xT4fDronGq9B062kmJ/2ua6TpFmhCRuPrehIVteVA0yAAWHB21szSA5ZeKfp1lBrZ
HbvdQVzwWO45iRH+xsUE1mnSGZY0RnKdkAUOADfE0UxQfK9BtorgwpLSXJDWzym1UBoImsiB9UeCIcXcr/31l7vJpDN6nX06zmuUf2mrAaftu5vPk/xz6OSf
p5PXTULOcLwsCfH7I1thhyduOxNyOhxnQnzP9vaRpSUyz+/5CutOPGcfVpawXc/P5J6MciO73fZYffZPR24j16nAsyLXlEYkRZ/ILbrkETi1SQ0yEz8InYaY
alAcAqQJMZahhvi0xqwR4BN9t74IyN+NiPerb5o9V6FYSdqE+BBGGuKcc+Zz0Wz7B6VG0fZVvNyjl1gVAZcY3zSqNSzF1niVwPGtnDwdExLNlAsGQYaXJCYS
qTl+TUgT/iul2v6c00DwlC8k+kqRj2mzI6d0Js3oMxrBRq8bdYdo0jx6/gX5nDUKHJEbHQJnG7NGIYRpu/AerySOmq3CEStCPmIZNhpytRaBtnGphGBaEsbR
eE7StBH8Waw1kz5gyOzNkXXO1pEOEZJeN0I+Ys6LkBG/HoY4SprtonFYBP2eXsNJweiCy2b9uH6G1TNsLI73R9QXSuQPJqc/6TI0B6OaWQm9hFZqn6qHND6o
HjIKBfG5Hj7lengKN5bGvFCugnsB/9HaN8Kr+ILAOX8ufc+l77n0PaHStzcjfWfB04tb3kZuW8T7rjHa1zQuKGNXcs3Ix1SvkynYOZ/A7P1oPp7x7frZJISv
mlktIxaQS4GzQSS4/IvK8CrECehkWyUJy1TTZTeKEp5CG27pU/VKldflr7kouDxb5OmvoXQ+LM/5PF/ntM0LM0O3ckvqtpS+tSY4SvSxzHBOHssMO2c8kh22
d6AdNfv2XXbkI6UwU5dDuBpCvgNtup3cOjiemJG5CtNSkG/D+enFeBriOdkEuX2YV23n2NHR++fBUbCj7zyWHceI8qIh7qGGmM/DQ4d5e1+YZ5XGUDQUbWys
JCxGt2C41/EsFOBkYC2gB4OvUQLyUlVgMVvGAyuQonxMjEXocOeXXF/j0ZLj26ZltW6vKXcZbSJSOcJpmBNnq8reZbHBVR3PVVvysL5qPbQVTs/+Wa3InwwR
ThYLEkhjlBemSqLzGVO+5ytJxFU4v0UzthKXGLzj5sdxTlO4Ena2DwIyubs5qXplMWem8t8tDAksW4hZEuJNXe3V55ucrnoidvqXd8Fg8v1wyUcP5TvnX/Rd
Q65+9t3j+m6TO0hMnHnFEQF0RQIjlRwGFhcy5FDukpAGEwHNlMlE8AKCZKYcgJj6C73yDLkpFc6tPjl/RSyDhk5e0iUSFIqwDAUhF3Lj7++TaneM1/osgW2E
VDJk1RfKQ4nBPTNyQ9hUJfOu2iYLhdviVM27Gr4mYEvDem6dLSf/217UPbQXPUbzo5ngHrOHc5t6uMJFrP9Y1h75Mt85cNs63gNe5hMsQ6R+wX2KioARq2K+
uq9P+SWcO7R78YEgm/zW26T23eAMfNSrWqVkKxE/Swd8H5IGY4xb9DRfjxRiraaxrcbaMQx5gFjzDKFmON+HRZoaM9WLrDmNCm9B1UDlP9vUDWj2DTQckQVe
MZm2NqPkTgo83P7vDbDCxI7h7Yu/AVBLAwQUAAAACADNcRhdjGo9mYsFAAAfEQAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbI1YbVPbOBD+KzuZueFu
LjjECS9tgRloj5Y52mYIbT8Lex3rkCVXkhvy729XtkNuLhH5ANiyXvZ59Ow+EudLY59ciejhuVLaXQxK7+u3o5HLSqyES0yNmr4UxlbC06tdjFxtUeRhUKVG
6dHRyagSUg8uz0PbzF6em8YrqXFmwTVVJezqGpVZXgzGg77hXi5Kzw2jy/NaLHCO/ls9s/Q2Ws+Sywq1k0aDxeJicDV+e51OeUDo8V3i0m08gyvN8qOV+R2t
TECOBsDgHo154s+3OTdRf1SYeZ5U0J9f+B6V4rkpsp/dMoN1FDxw87lf7ybQQfAehcP3Rv2QuS8vBmcDyLEQjfL3ZvkJO4jHPF9mlAu/Ydn2TdMBZI3zpuoG
UwSV1O1f8dxRszFgfDTdMSLtRqQh8HalEOYH4cXluTVLsNybpuOHgDWMpuik5o2ae0tfJY3zlzNr/iGGgLYcfInwAx/hEP56zlDBbVUb6+EBq1oJj+cjTwvy
sFFGP7TQerXJABh8etbHvKZjHcYkhJH+LwyeK/S4bntMdgR6jxREhi5EadFLizkcdCEW1lTweTY7gFr4MoEbqRRIHfr60iJCYMjBIytzCE2tjMjD50Iq6mNA
6B5vT8mckiJvFK+WGZsPwTYavgslcyJjyIN1aGqHJRF6jtcsHAeM0x0YP1F3CqVxMa5PWq7H091cn+zgul1lnKwhBlb+u1a7FyfRvfgiKgzk1e08QyD11ML5
tvX2A5iCCcVn6bzUCzA6cCxbgqXm5yhjp6+DPI2CTBO4R2cay5LZCfM0CvOOogf8hXYFNVpHRYRgLlCjlTTcsG5K4UHUNQrrer09CPfULRlDePY6wrMowkmy
udQ2dGdRdF9pTzgWggaeJhr2AEjr1LQyDSyF9txUhQQJMPtOrkuOGMQ3r0N8E4U4TeBbm6l/rhNvG9A3UaDdSDaBpZWe9KCNL1mWhGpDxVw2nkm2K0BrDW2o
0Dnl/iHneFyt46PXkXKfCNTjpKsj2/B1Y3eWxkYH9T0KT7vi3oEAJeziZZNof58ItoCCjJNMpCESonjSl7DTaMW6Q8qHPDbVZA9qdnlDu8a8FDnVenIk5bay
EzcOtnwX5Fy0rpDAR4sr2lLSBefxkpixSMVKkM8RTYdk6wrbWlaRY5A1Yle6SDRR3qZ7gJ1GwXJKUwHdinMaxXlFwm20/Nkg7T75XICs5BPC7+MhpEOYcPn6
cT2HzOTo/qDcckSr0WrFiUDTPbVk0AuKrARKErRRuCcvqE6iMnngdHObhYPTD/WC+hHbmaSAwhGEYo4tuIctjOO+MG8PpKHibVdT3BPokIVuCHlj24rCNQKL
gm2N38K8rS2UwkFWSpVbOiewwORCGz60hLJilKLHpo7yu4dHjOMmcSO1dCVw4dyONu4RH+nIDHMvCF2tGgcfOthcKsPE0tH2WeqVt2ewjXIKmVBUG0RcQnt4
xDhuEncsXL+qdwCMe8NXVn8RsBx6c+gCVALlmprzHfME7sQiwLSGDJBOgZgsEjiY3Mz/TPODri5QBeHyGnKITYVFMInhTvdwjDTuGL/Be8MFa7slpnHLuKpr
JVmLpNFulhx+ObI7vtyF5AyiTuCL8S9ssMb7s+NXOjgEMUStJH2xkvQVK5GV9C421R5Wksat5GFX1qdxD/kbsab97ZXdaD4hsdZzkzV0bWVqiI7DvkMQQIaS
y9w7cMS1by3Zrj05ztr09QtVGjeSO5nxdVovtsKNW8kDZ3F74qXDT5ADg+2O7xoxZ7X3QpgpOuorXi5DMs/N0yFczW5ZVKGfYM1lL2WTrgG46MooXXnsdkZG
G3db/tfBZ6JR0nlHYUGxHyWndKWyLTftizd1uO4+Gk+8hccS6QRhuQN9L4zx/QvfoNf/E7n8F1BLAwQUAAAACADNcRhdMylC8AgEAABgDAAAGAAAAHhsL3dv
cmtzaGVldHMvc2hlZXQyLnhtbM1X227jNhD9lYGAAlsgayW+I7ANxPa2DVCnxmazQZ8KWhpbbChSJal13K/vkJK1TiKzQB6KviTUkHM5hzPD8WSv9JPJEC08
50KaaZRZW1zHsUkyzJnpqAIl7WyVzpmlT72LTaGRpV4pF3H38nIY54zLaDbxsrWeTVRpBZe41mDKPGf6MEeh9tPoKjoKPvNdZp0gnk0KtsN7tA/FWtNX3FhJ
eY7ScCVB43Ya3VxdL8buvD/wlePenKzBIdko9eQ+btNpdBk5yxLhcF8I7n2BVcWvuLULFILsdSNgieXfcE3HptFGWatyt09RWmZJtNXqb5TeJwqksxRL8eZw
ZaQ26iD+VccbNXBcUKfrY+Q/eV6Jpw0zuFDikac2m0bjCFLcslLYz2r/C9ZcDZy9RAnj/8K+OtsjGElpKJpamSLIuaz+s+ea4xOF/vCMQrdW6L5SGJ7z0KsV
eh5oFZmHtWSWzSZa7UG702TNLTw302hA9zCNuHQZcm817XLSszNKAcvlbhJbMuZEcVIrzsOKX5kosUVtEVa7UxbNS7WYQm7i7jZxd72d4Rk7a63+pNyAO5a3
RTGvtEde21XJ9/iqnfEZuw8GU1BSHGCfoYQvTO+oTI/ebpfADWwEk0+dAIheA6IXBPHGehuS3lkkvSCSNTMWgUnAZ27cHUNu0oP8o6i9/fxAaKwCnhdKW+DS
rW0H5g4dJNRu6KaAgaQyr3VCmPsN5n4Q8xILoQ7UY2wb2P4J2POXTo0RbIbwiJu2DOyHeXlrBGh5FP9WoGau55gQ2kGDdhBEu/A0woob427gC7IcVphvUJs2
+IMg/N9fV04FdxCEe5Om1MmNKnVC10l37DAfc8C6eFhJTYYgJ0yIQwj0sAE9DIK+KQqqnx9gofJCoG2tz+F7kA6DSB81tx7ajvBS5hpI6gDSWKN7Ld0V4Hbr
0p1tLdL9M/NEJzXW+Z6G4I8a+KMg/FuZiDJFWPoCupEpfPI+22gYvYeGUZAG6ujumu+Uz3BGr6ouE1tq/OjbGsVkQZfyGiS1TgJP8WWccl4n2eECpAI3haSl
QEjpQQkRMm4IGQcJWXFZOirWxPiSHdp4GAd56I8v23gY/3snJyISJb8h3fiyrOoaPqTsYH50W3kVWAfoLJCTijEJ448Z1YsfblzK0PlWGuKTZ9dxRU8iT6ve
QV5LaauX4OXWcVCZD8AXQab2S62KpdpLNz55wa0sSruiJKYZrRF+0lrpUyGVq9r7bl2NWYeC5IJ6Pfl0o2Mp2NUsohy6uFPRJG5Ek/hlRGcj7P8H4bV044u3
rfhd4Y/+J+y+Ephq9l7R288pUQRNs+SwM6I2rquZs/qgsdl7roZev8zoJwBqd4D2t4pGqfrDTYLNj4rZP1BLAwQUAAAACADNcRhdyQtMu+gBAAB7BAAAGAAA
AHhsL3dvcmtzaGVldHMvc2hlZXQzLnhtbIVU7Y7TMBB8FcsPcG5TesApiUSDEEiAqlYHv91001jnj2BvyR1Pz9pJQ+9o4U/i9c5MZtfr5L3zD6EFQPZotA0F
bxG7OyFC3YKR4cZ1YCnTOG8kUugPInQe5D6RjBbZbHYrjFSWl3naW/syd0fUysLas3A0RvqnFWjXF3zOTxsbdWgxbogy7+QBtoD33dpTJCaVvTJgg3KWeWgK
/m5+Vy0jPgG+KejD2ZrFSnbOPcTg077gMx6VLbCnbadV+hZD132GBivQmvQyzmSN6iesCVbwnUN0JubJJUqkrca7X2DTN0EDYclL9xd4EBlFY4k/Rr98Kiea
Ol+fnH9IfaU+7WSAyunvao9twd9wtodGHjVuXP8Rxl6l4munQ3qyfsBmt5zVx0BuRjI5MMoOb/k49viMsHh1hZCNhOwFIbtGWIyERSp0cJbKei9Rlrl3PfMR
TWpxkXpT8CWdQ8GVjROyRU9ZRTwsNxDc0dfAvkoDuUCSjAlRj/TVv+krOnu508BOOhckqv84cPoFS1ANUyHZVEiWZN4mmTj4fzxezVSXMs/kF5P84qr81Ux1
KfNMfjnJLxN0PruAFWfHF6/lF+kPygamadDpSt28Jq4fxnEI6EaleRjuQ1q29HcAHwGUb5zDUxCHZPrflL8BUEsDBBQAAAAIAM5xGF2aPURd9SwAADoTAgAY
AAAAeGwvd29ya3NoZWV0cy9zaGVldDQueG1std1bj93Hnd/rt0IICJB9M2KvPhu2gLjKq06ZbGNmkn3dlloWYYqtIVujmbz63aTIWrUqVY+MAHNjk/3w24cf
m7I/trT+v//l6f3fPvzw+Pj86t9/fPvuwx+++uH5+affff31h29/ePzx4cM/PP30+O5Fvn96/+PD88tP3//16w8/vX98+O7T6Me3Xx9ev775+seHN++++ub3
n9725/ff/P7p5+e3b949/vn9qw8///jjw/v/+OPj26df/vDVxVdf3vBPb/76w/PHN3z9ze9/evjr4z8/Pv/Pn/78/uVnX/f38t2bHx/ffXjz9O7V+8fv//DV
f7v4Xbu8v/+4+PRL/tebx18+DD9+9fFr+cvT098+/qR894evXn/18X2/e3z17//809s3Lx/t8qtX//H5hy+fyfPTT//98fvn8Pj27R++ioevXj18+/zm3x7/
/LL4w1d/eXp+fvrx02f58jk/Pzy/vO3790//+/Hdp4//+Pbx5Re/fGY/ffrVL+/q8y/99X38+j7/28cv918/f+6r2a8f5OMn8X81/Hs+5Nf9POOPv5zt+Om3
9eW36S8PHx7D09v/7813zz/84au7r1599/j9w89vn//p6Zf8+Pm36vrj+/v26e2HT//66pdff+3F669effvzh5dP5/P45TP48c27X//94d8//xaPg+vN4PB5
cJgGV7uPcPl5cDl/hMNmcPV5cPX3fkrXnwfXf+9HuPk8uJkHV5vB7efB7d/7Ee4+D+7mwd1mcP95cD8NDje737jXX37nXk+Ty91XcdF/s3/9pvv1u+TTt1h8
eH745vfvn3559f7Tr//4rXTon2v/5nr5o/rtx1/x6Rv413f48uY37z7+ZeSfn9+/8JuX9/j8zb88fPjbqxJ///Xzy4f5+Kavv/28/ONvLP/88P7x3fOr/TsI
f8+H/h8PPz4utvE3tv/8/PD+ebH702/s4s/vHz79uf+v3z38x4f/Z/Eejr/xHo5v3r358MNimH5j+KfvX/6i//zqv/7w9PP75UfOv/EO/sur8PTjT28fn1cH
K7/1u/X+8bvHbx8/fHh6/2Exr78x/6fHDy+f9st+sW2/sf0fT8/z7uuX79/+TXzo36uHT+/o/tP7+fiff6fvxa2ErcRf5eMf/Jn+tB0d96O0HeXPo8v/k8p2
VLfSVnJ2sct+scvtxbYSthIv9xfbjo77UdqO8uX+YttR3UpbydnFrvrFrrYX20rYSrzaX2w7Ou5HaTvKV/uLbUd1K20lZxe77he73l5sK2Er8Xp/se3ouB+l
7Shf7y+2HdWttJWcXeymX+zm1498tTjZnsKe4me6XhxtvzruV2m/ynsqe6p7aks6u9xtv9zt/nJ7CnuKt/vL7VfH/SrtV3lPZU91T21JZ5e765e7219uT2FP
8W5/uf3quF+l/Srvqeyp7qkt6exy9/1y9/vL7SnsKd7vL7dfHfertF/lPZU91T21JZ1d7uL1KS5e728HC7D4xVbnw+6IXcIuwwqswtrazq84JNoFrri3AItf
bHnF/e6IXcIuwwqswtrazq94ioeLA664twCLX2x5xf3uiF3CLsMKrMLa2s6veAqKi0tccW8BFr/Y8or73RG7hF2GFViFtbWdX/EUGRdXuOLeAix+seUV97sj
dgm7DCuwCmtrO7/iKTwurnHFvQVY/GLLK+53R+wSdhlWYBXW1nZ+xVOMXKBGYAEWLxAk2B2xS9hlWIFVWFvb+RVPYXKBMoEFWLxAnGB3xC5hl2EFVmFtbedX
PEXKBSoFFmDxAqGC3RG7hF2GFViFtbWdX/EULBcoFliAxQtEC3ZH7BJ2GVZgFdbWdv6/NZ/a5YB2gQVYPKBdsDtil7DLsAKrsLa28yue2uWAdoEFWDygXbA7
Ypewy7ACq7C2tvMrDv/HB9oFFmDxgHbB7ohdwi7DCqzC2trOr3hqlwPaBRZg8YB2we6IXcIuwwqswtrazq94apcD2gUWYPGLLa+IdsEuYZdhBVZhbW3nVzy1
ywHtAguweEC7YHfELmGXYQVWYW1t51c8tcsB7QILsHhAu2B3xC5hl2EFVmFtbedXPLXLAe0CC7B4QLtgd8QuYZdhBVZhbW3nVzy1ywHtAguweEC7YHfELmGX
YQVWYW1t51c8tcsB7QILsHhAu2B3xC5hl2EFVmFtbed/18epXS7RLrAAi5doF+yO2CXsMqzAKqyt7fyKp3a5RLvAAixeol2wO2KXsMuwAquwtrbzK57a5RLt
AguweIl2we6IXcIuwwqswtrazq84/I1caBdYgMVLtAt2R+wSdhlWYBXW1nZ+xVO7XKJdYAEWv9jyimgX7BJ2GVZgFdbWdn7FU7tcol1gARYv0S7YHbFL2GVY
gVVYW9v5FU/tcol2gQVYvES7YHfELmGXYQVWYW1t51c8tcsl2gUWYPES7YLdEbuEXYYVWIW1tZ1f8dQul2gXWIDFS7QLdkfsEnYZVmAV1tZ2fsVTu1yiXWAB
Fi/RLtgdsUvYZViBVVhb2/nff31qlyu0CyzA4hXaBbsjdgm7DCuwCmtrO7/iqV2u0C6wAItXaBfsjtgl7DKswCqsre38iqd2uUK7wAIsXqFdsDtil7DLsAKr
sLa28yue2uUK7QILsHiFdsHuiF3CLsMKrMLa2s6vOPyDKWgXWIDFL7a8ItoFu4RdhhVYhbW1nV/x1C5XaBdYgMUrtAt2R+wSdhlWYBXW1nZ+xVO7XKFdYAEW
r9Au2B2xS9hlWIFVWFvb+RVP7XKFdoEFWLxCu2B3xC5hl2EFVmFtbedXPLXLFdoFFmDxCu2C3RG7hF2GFViFtbWdX/HULldoF1iAxSu0C3ZH7BJ2GVZgFdbW
dv5PQp7a5RrtAguweI12we6IXcIuwwqswtrazq94apdrtAsswOI12gW7I3YJuwwrsAprazu/4qldrtEusACL12gX7I7YJewyrMAqrK3t/IqndrlGu8ACLF6j
XbA7Ypewy7ACq7C2tvMrntrlGu0CC7D4xZZXRLtgl7DLsAKrsLa28ysO/6A92gUWYPEa7YLdEbuEXYYVWIW1tZ1f8dQu12gXWIDFa7QLdkfsEnYZVmAV1tZ2
fsVTu1yjXWABFq/RLtgdsUvYZViBVVhb2/kVT+1yjXaBBVi8Rrtgd8QuYZdhBVZhbW3nVzy1yzXaBRZg8Rrtgt0Ru4RdhhVYhbW1nb8myaldbtAusACLN2gX
7I7YJewyrMAqrK3t/IqndrlBu8ACLN6gXbA7Ypewy7ACq7C2tvMrntrlBu0CC7B4g3bB7ohdwi7DCqzC2trOr3hqlxu0CyzA4g3aBbsjdgm7DCuwCmtrO7/i
qV1u0C6wAItfbHlFtAt2CbsMK7AKa2s7v+KpXW7QLrAAizdoF+yO2CXsMqzAKqyt7fyKwwuH6ZXD9NJheu0wvXiYXj1MLx+m1w/TC4jpFcT0EmK/3S43p3a5
QbvAAizeoF2wO2KXsMuwAquwtrbzK57a5QbtAguweIN2we6IXcIuwwqswtrazq94apcbtAsswOIN2gW7I3YJuwwrsAprazt/VcBTu9yiXWABFm/RLtgdsUvY
ZViBVVhb2/kVT+1yi3aBBVi8Rbtgd8QuYZdhBVZhbW3nVzy1yy3aBRZg8Rbtgt0Ru4RdhhVYhbW1nV/x1C63aBdYgMVbtAt2R+wSdhlWYBXW1nZ+xVO73KJd
YAEWv9jyimgX7BJ2GVZgFdbWdn7FU7vcol1gARZv0S7YHbFL2GVYgVVYW9v5FU/tcot2gQVYvEW7YHfELmGXYQVWYW1t51ccXgBZr4Csl0DWayDrRZD1Ksh6
GWS9DrJeCFmvhKyXQv7tdrk9tcst2gUWYPEW7YLdEbuEXYYVWIW1tZ1f8dQut2gXWIDFW7QLdkfsEnYZVmAV1tZ2/rrcp3a5Q7vAAizeoV2wO2KXsMuwAquw
trbzK57a5Q7tAguweId2we6IXcIuwwqswtrazq94apc7tAsswOId2gW7I3YJuwwrsAprazu/4qld7tAusACLd2gX7I7YJewyrMAqrK3t/IqndrlDu8ACLH6x
5RXRLtgl7DKswCqsre38iqd2uUO7wAIs3qFdsDtil7DLsAKrsLa28yue2uUO7QILsHiHdsHuiF3CLsMKrMLa2s6veGqXO7QLLMDiHdoFuyN2CbsMK7AKa2s7
v+LwIBc9yUWPctGzXPQwFz3NRY9z0fNc9EAXPdFFj3T57Xa5O7XLHdoFFmDxDu2C3RG7hF2GFViFtbWdPxnn1C73aBdYgMV7tAt2R+wSdhlWYBXW1nZ+xVO7
3KNdYAEW79Eu2B2xS9hlWIFVWFvb+RVP7XKPdoEFWLxHu2B3xC5hl2EFVmFtbedXPLXLPdoFFmDxHu2C3RG7hF2GFViFtbWdX/HULvdoF1iAxS+2vCLaBbuE
XYYVWIW1tZ1f8dQu92gXWIDFe7QLdkfsEnYZVmAV1tZ2fsVTu9yjXWABFu/RLtgdsUvYZViBVVhb2/kVT+1yj3aBBVi8R7tgd8QuYZdhBVZhbW3nVzy1yz3a
BRZg8R7tgt0Ru4RdhhVYhbW1nV9xeCClnkipR1LqmZR6KKWeSqnHUuq5lHowpZ5MqUdT/j3PphwfTsmnU/LxlHw+JR9QySdU8hGVfEYlH1LJp1TyMZV/x3Mq
Xw8PqnytJ1UCgzB2XF9UT6vEMmmZhUVYhW2D00WHh1a+1lMrgUEYO64vqidXYpm0zMIirMK2wemiwwMsX+sJlsAgjB3XF9VTLLFMWmZhEVZh2+B00eFhlq9R
O8IgjB3XF0XxaJm0zMIirMK2wemiw4MtX6N8hEEYO64vivrRMmmZhUVYhW2D00WHh1y+1lMugUEYO64vqiddYpm0zMIirMK2wemiwwMvX+uJl8AgjB3XF9VT
L7FMWmZhEVZh2+B00eHhl6/19EtgEMaO64vqCZhYJi2zsAirsG1wuujwIMzXehImMAhjx/VF9TRMLJOWWViEVdg2OD2JfmgmPNX+j8IgjB2XF8XyqGXSMguL
sArbBqeLDs2EJ9z/URiEseP6omomLJOWWViEVdg2OF10aCY87f6PwiCMHdcXVTNhmbTMwiKswrbB6aJDM62fbv/lomomYOy4vqiaCcukZRYWYRW2DU4XHZrp
Qs0EDMLYcX1RNROWScssLMIqbBucLjo004WaCRiEseP6omomLJOWWViEVdg2OF10aKYLNRMwCGPH9UXVTFgmLbOwCKuwbXC66NBMF2omYBDGjuuLqpmwTFpm
YRFWYdvgdNGhmS7UTMAgjB3XF1UzYZm0zMIirMK2wemiQzNdqJmAQRg7ri+qZsIyaZmFRViFbYPnFz0MzXRQMwGDMHZcXhTLo5ZJyywswipsG5wuOjTTQc0E
DMLYcX1RNROWScssLMIqbBucLjo000HNBAzC2HF9UTUTlknLLCzCKmwbnC46NNNBzQQMwthxfVE1E5ZJyywswipsG5wuOjTTQc0EDMLYcX1RNROWScssLMIq
bBucLjo000HNBAzC2HF9UTUTlknLLCzCKmwbnC46NNNBzQQMwthxfVE1E5ZJyywswipsG5wuOjTTQc0EDMLYcX1RNROWScssLMIqbBucLjo000HNBAzC2HF9
UTUTlknLLCzCKmwbnC46NNNBzQQMwthxfVE1E5ZJyywswipsGzy/6OXQTJdqJmAQxo7Li2J51DJpmYVFWIVtg9NFh2a6VDMBgzB2XF9UzYRl0jILi7AK2wan
iw7NdKlmAgZh7Li+qJoJy6RlFhZhFbYNThcdmulSzQQMwthxfVE1E5ZJyywswipsG5wuOjTTpZoJGISx4/qiaiYsk5ZZWIRV2DY4XXRopks1EzAIY8f1RdVM
WCYts7AIq7BtcLro0EyXaiZgEMaO64uqmbBMWmZhEVZh2+B00aGZLtVMwCCMHdcXVTNhmbTMwiKswrbB6aJDM12qmYBBGDuuL6pmwjJpmYVFWIVtg9NFh2a6
VDMBgzB2XF9UzYRl0jILi7AK2wbPL3o1NNOVmgkYhLHj8qJYHrVMWmZhEVZh2+B00aGZrtRMwCCMHdcXVTNhmbTMwiKswrbB6aJDM12pmYBBGDuuL6pmwjJp
mYVFWIVtg9NFh2a6UjMBgzB2XF9UzYRl0jILi7AK2waniw7NdKVmAgZh7Li+qJoJy6RlFhZhFbYNThcdmulKzQQMwthxfVE1E5ZJyywswipsG5wuOjTTlZoJ
GISx4/qiaiYsk5ZZWIRV2DY4XXRopis1EzAIY8f1RdVMWCYts7AIq7BtcLro0ExXaiZgEMaO64uqmbBMWmZhEVZh2+B00aGZrtRMwCCMHdcXVTNhmbTMwiKs
wrbB84teD810rWYCBmHsuLwolkctk5ZZWIRV2DY4XXRopms1EzAIY8f1RdVMWCYts7AIq7BtcLro0EzXaiZgEMaO64uqmbBMWmZhEVZh2+B00aGZrtVMwCCM
HdcXVTNhmbTMwiKswrbB6aJDM12rmYBBGDuuL6pmwjJpmYVFWIVtg9NFh2a6VjMBgzB2XF9UzYRl0jILi7AK2waniw7NdK1mAgZh7Li+qJoJy6RlFhZhFbYN
ThcdmulazQQMwthxfVE1E5ZJyywswipsG5wuOjTTtZoJGISx4/qiaiYsk5ZZWIRV2DY4XXRopms1EzAIY8f1RdVMWCYts7AIq7Bt8PyiN0Mz3aiZgEEYOy4v
iuVRy6RlFhZhFbYNThcdmulGzQQMwthxfVE1E5ZJyywswipsG5wuOjTTjZoJGISx4/qiaiYsk5ZZWIRV2DY4XXRophs1EzAIY8f1RdVMWCYts7AIq7BtcLro
0Ew3aiZgEMaO64uqmbBMWmZhEVZh2+B00aGZbtRMwCCMHdcXVTNhmbTMwiKswrbB6aJDM92omYBBGDuuL6pmwjJpmYVFWIVtg9NFh2a6UTMBgzB2XF9UzYRl
0jILi7AK2waniw7NdKNmAgZh7Li+qJoJy6RlFhZhFbYNThcdmulGzQQMwthxfVE1E5ZJyywswipsGzy/6O3QTLdqJmAQxo7Li2J51DJpmYVFWIVtg9NFh2a6
VTMBgzB2XF9UzYRl0jILi7AK2waniw7NdKtmAgZh7Li+qJoJy6RlFhZhFbYNThcdmulWzQQMwthxfVE1E5ZJyywswipsG5wuOjTTrZoJGISx4/qiaiYsk5ZZ
WIRV2DY4XXRopls1EzAIY8f1RdVMWCYts7AIq7BtcLro0Ey3aiZgEMaO64uqmbBMWmZhEVZh2+B00aGZbtVMwCCMHdcXVTNhmbTMwiKswrbB6aJDM92qmYBB
GDuuL6pmwjJpmYVFWIVtg9NFh2a6VTMBgzB2XF9UzYRl0jILi7AK2wbPL3o3NNOdmgkYhLHj8qJYHrVMWmZhEVZh2+B00aGZ7tRMwCCMHdcXVTNhmbTMwiKs
wrbB6aJDM92pmYBBGDuuL6pmwjJpmYVFWIVtg9NFh2a6UzMBgzB2XF9UzYRl0jILi7AK2waniw7NdKdmAgZh7Li+qJoJy6RlFhZhFbYNThcdmulOzQQMwthx
fVE1E5ZJyywswipsG5wuOjTTnZoJGISx4/qiaiYsk5ZZWIRV2DY4XXRopjs1EzAIY8f1RdVMWCYts7AIq7BtcLro0Ex3aiZgEMaO64uqmbBMWmZhEVZh2+B0
0aGZ7tRMwCCMHdcXVTNhmbTMwiKswrbB84veD810r2YCBmHsuLwolkctk5ZZWIRV2DY4XXRopns1EzAIY8f1RdVMWCYts7AIq7BtcLro0Ez3aiZgEMaO64uq
mbBMWmZhEVZh2+B00aGZ7tVMwCCMHdcXVTNhmbTMwiKswrbB6aJDM92rmYBBGDuuL6pmwjJpmYVFWIVtg9NFh2a6VzMBgzB2XF9UzYRl0jILi7AK2waniw7N
dK9mAgZh7Li+qJoJy6RlFhZhFbYNThcdmulezQQMwthxfVE1E5ZJyywswipsG5wuOjTTvZoJGISx4/qiaiYsk5ZZWIRV2DY4XXRopns1EzAIY8f1RdVMWCYt
s7AIq7Bt8Oyih9enZvr44+1FhUEYO64uquVRy6RlFhZhFbYNThe9GC6KZhIGYey4viiaScukZRYWYRW2DU4XPQwXRTMJgzB2XF8UzaRl0jILi7AK2wani14O
F0UzCYMwdlxfFM2kZdIyC4uwCtsGp4teDRdFMwmDMHZcXxTNpGXSMguLsArbBqeLXg8XRTMJgzB2XF8UzaRl0jILi7AK2wani94MF0UzCYMwdlxfFM2kZdIy
C4uwCtsGp4veDhdFMwmDMHZcXxTNpGXSMguLsArbBqeL3g0XRTMJgzB2XF8UzaRl0jILi7AK2wani94PF0UzCYMwdlxfFM2kZdIyC4uwCtsGzy96MTTThZoJ
GISx4/KiWB61TFpmYRFWYdvgdNGhmS7UTMAgjB3XF1UzYZm0zMIirMK2wemiQzNdqJmAQRg7ri+qZsIyaZmFRViFbYPTRYdmulAzAYMwdlxfVM2EZdIyC4uw
CtsGp4sOzXShZgIGYey4vqiaCcukZRYWYRW2DU4XHZrpQs0EDMLYcX1RNROWScssLMIqbBucLjo004WaCRiEseP6omomLJOWWViEVdg2OF10aKYLNRMwCGPH
9UXVTFgmLbOwCKuwbXC66NBMF2omYBDGjuuLqpmwTFpmYRFWYdvgdNGhmS7UTMAgjB3XF1UzYZm0zMIirMK2wfOLHoZmOqiZgEEYOy4viuVRy6RlFhZhFbYN
ThcdmumgZgIGYey4vqiaCcukZRYWYRW2DU4XHZrpoGYCBmHsuL6omgnLpGUWFmEVtg1OFx2a6aBmAgZh7Li+qJoJy6RlFhZhFbYNThcdmumgZgIGYey4vqia
CcukZRYWYRW2DU4XHZrpoGYCBmHsuL6omgnLpGUWFmEVtg1OFx2a6aBmAgZh7Li+qJoJy6RlFhZhFbYNThcdmumgZgIGYey4vqiaCcukZRYWYRW2DU4XHZrp
oGYCBmHsuL6omgnLpGUWFmEVtg1OFx2a6aBmAgZh7Li+qJoJy6RlFhZhFbYNnl/0cmimSzUTMAhjx+VFsTxqmbTMwiKswrbB6aJDM12qmYBBGDuuL6pmwjJp
mYVFWIVtg9NFh2a6VDMBgzB2XF9UzYRl0jILi7AK2waniw7NdKlmAgZh7Li+qJoJy6RlFhZhFbYNThcdmulSzQQMwthxfVE1E5ZJyywswipsG5wuOjTTpZoJ
GISx4/qiaiYsk5ZZWIRV2DY4XXRopks1EzAIY8f1RdVMWCYts7AIq7BtcLro0EyXaiZgEMaO64uqmbBMWmZhEVZh2+B00aGZLtVMwCCMHdcXVTNhmbTMwiKs
wrbB6aJDM12qmYBBGDuuL6pmwjJpmYVFWIVtg+cXvRqa6UrNBAzC2HF5USyPWiYts7AIq7BtcLro0ExXaiZgEMaO64uqmbBMWmZhEVZh2+B00aGZrtRMwCCM
HdcXVTNhmbTMwiKswrbB6aJDM12pmYBBGDuuL6pmwjJpmYVFWIVtg9NFh2a6UjMBgzB2XF9UzYRl0jILi7AK2waniw7NdKVmAgZh7Li+qJoJy6RlFhZhFbYN
ThcdmulKzQQMwthxfVE1E5ZJyywswipsG5wuOjTTlZoJGISx4/qiaiYsk5ZZWIRV2DY4XXRopis1EzAIY8f1RdVMWCYts7AIq7BtcLro0ExXaiZgEMaO64uq
mbBMWmZhEVZh2+D5Ra+HZrpWMwGDMHZcXhTLo5ZJyywswipsG5wuOjTTtZoJGISx4/qiaiYsk5ZZWIRV2DY4XXRopms1EzAIY8f1RdVMWCYts7AIq7BtcLro
0EzXaiZgEMaO64uqmbBMWmZhEVZh2+B00aGZrtVMwCCMHdcXVTNhmbTMwiKswrbB6aJDM12rmYBBGDuuL6pmwjJpmYVFWIVtg9NFh2a6VjMBgzB2XF9UzYRl
0jILi7AK2waniw7NdK1mAgZh7Li+qJoJy6RlFhZhFbYNThcdmulazQQMwthxfVE1E5ZJyywswipsG5wuOjTTtZoJGISx4/qiaiYsk5ZZWIRV2DZ4ftGboZlu
1EzAIIwdlxfF8qhl0jILi7AK2waniw7NdKNmAgZh7Li+qJoJy6RlFhZhFbYNThcdmulGzQQMwthxfVE1E5ZJyywswipsG5wuOjTTjZoJGISx4/qiaiYsk5ZZ
WIRV2DY4XXRophs1EzAIY8f1RdVMWCYts7AIq7BtcLro0Ew3aiZgEMaO64uqmbBMWmZhEVZh2+B00aGZbtRMwCCMHdcXVTNhmbTMwiKswrbB6aJDM92omYBB
GDuuL6pmwjJpmYVFWIVtg9NFh2a6UTMBgzB2XF9UzYRl0jILi7AK2waniw7NdKNmAgZh7Li+qJoJy6RlFhZhFbYNnl/0dmimWzUTMAhjx+VFsTxqmbTMwiKs
wrbB6aJDM92qmYBBGDuuL6pmwjJpmYVFWIVtg9NFh2a6VTMBgzB2XF9UzYRl0jILi7AK2waniw7NdKtmAgZh7Li+qJoJy6RlFhZhFbYNThcdmulWzQQMwthx
fVE1E5ZJyywswipsG5wuOjTTrZoJGISx4/qiaiYsk5ZZWIRV2DY4XXRopls1EzAIY8f1RdVMWCYts7AIq7BtcLro0Ey3aiZgEMaO64uqmbBMWmZhEVZh2+B0
0aGZbtVMwCCMHdcXVTNhmbTMwiKswrbB6aJDM92qmYBBGDuuL6pmwjJpmYVFWIVtg+cXvRua6U7NBAzC2HF5USyPWiYts7AIq7BtcLro0Ex3aiZgEMaO64uq
mbBMWmZhEVZh2+B00aGZ7tRMwCCMHdcXVTNhmbTMwiKswrbB6aJDM92pmYBBGDuuL6pmwjJpmYVFWIVtg9NFh2a6UzMBgzB2XF9UzYRl0jILi7AK2waniw7N
dKdmAgZh7Li+qJoJy6RlFhZhFbYNThcdmulOzQQMwthxfVE1E5ZJyywswipsG5wuOjTTnZoJGISx4/qiaiYsk5ZZWIRV2DY4XXRopjs1EzAIY8f1RdVMWCYt
s7AIq7BtcLro0Ex3aiZgEMaO64uqmbBMWmZhEVZh2+D5Re+HZrpXMwGDMHZcXhTLo5ZJyywswipsG5wuOjTTvZoJGISx4/qiaiYsk5ZZWIRV2DY4XXRopns1
EzAIY8f1RdVMWCYts7AIq7BtcLro0Ez3aiZgEMaO64uqmbBMWmZhEVZh2+B00aGZ7tVMwCCMHdcXVTNhmbTMwiKswrbB6aJDM92rmYBBGDuuL6pmwjJpmYVF
WIVtg9NFh2a6VzMBgzB2XF9UzYRl0jILi7AK2waniw7NdK9mAgZh7Li+qJoJy6RlFhZhFbYNThcdmulezQQMwthxfVE1E5ZJyywswipsG5wuOjTTvZoJGISx
4/qiaiYsk5ZZWIRV2DZ4dtHL16dm+vjj7UWFQRg7ri6q5VHLpGUWFmEVtg1OF70YLopmEgZh7Li+KJpJy6RlFhZhFbYNThc9DBdFMwmDMHZcXxTNpGXSMguL
sArbBqeLXg4XRTMJgzB2XF8UzaRl0jILi7AK2wani14NF0UzCYMwdlxfFM2kZdIyC4uwCtsGp4teDxdFMwmDMHZcXxTNpGXSMguLsArbBqeL3gwXRTMJgzB2
XF8UzaRl0jILi7AK2wani94OF0UzCYMwdlxfFM2kZdIyC4uwCtsGp4veDRdFMwmDMHZcXxTNpGXSMguLsArbBqeL3g8XRTMJgzB2XF8UzaRl0jILi7AK2wbP
L3oxNNOFmgkYhLHj8qJYHrVMWmZhEVZh2+B00aGZLtRMwCCMHdcXVTNhmbTMwiKswrbB6aJDM12omYBBGDuuL6pmwjJpmYVFWIVtg9NFh2a6UDMBgzB2XF9U
zYRl0jILi7AK2waniw7NdKFmAgZh7Li+qJoJy6RlFhZhFbYNThcdmulCzQQMwthxfVE1E5ZJyywswipsG5wuOjTThZoJGISx4/qiaiYsk5ZZWIRV2DY4XXRo
pgs1EzAIY8f1RdVMWCYts7AIq7BtcLro0EwXaiZgEMaO64uqmbBMWmZhEVZh2+B00aGZLtRMwCCMHdcXVTNhmbTMwiKswrbB84sehmY6qJmAQRg7Li+K5VHL
pGUWFmEVtg1OFx2a6aBmAgZh7Li+qJoJy6RlFhZhFbYNThcdmumgZgIGYey4vqiaCcukZRYWYRW2DU4XHZrpoGYCBmHsuL6omgnLpGUWFmEVtg1OFx2a6aBm
AgZh7Li+qJoJy6RlFhZhFbYNThcdmumgZgIGYey4vqiaCcukZRYWYRW2DU4XHZrpoGYCBmHsuL6omgnLpGUWFmEVtg1OFx2a6aBmAgZh7Li+qJoJy6RlFhZh
FbYNThcdmumgZgIGYey4vqiaCcukZRYWYRW2DU4XHZrpoGYCBmHsuL6omgnLpGUWFmEVtg2eX/RyaKZLNRMwCGPH5UWxPGqZtMzCIqzCtsHpokMzXaqZgEEY
O64vqmbCMmmZhUVYhW2D00WHZrpUMwGDMHZcX1TNhGXSMguLsArbBqeLDs10qWYCBmHsuL6omgnLpGUWFmEVtg1OFx2a6VLNBAzC2HF9UTUTlknLLCzCKmwb
nC46NNOlmgkYhLHj+qJqJiyTlllYhFXYNjhddGimSzUTMAhjx/VF1UxYJi2zsAirsG1wuujQTJdqJmAQxo7ri6qZsExaZmERVmHb4HTRoZku1UzAIIwd1xdV
M2GZtMzCIqzCtsHpokMzXaqZgEEYO64vqmbCMmmZhUVYhW2D5xe9GprpSs0EDMLYcXlRLI9aJi2zsAirsG1wuujQTFdqJmAQxo7ri6qZsExaZmERVmHb4HTR
oZmu1EzAIIwd1xdVM2GZtMzCIqzCtsHpokMzXamZgEEYO64vqmbCMmmZhUVYhW2D00WHZrpSMwGDMHZcX1TNhGXSMguLsArbBqeLDs10pWYCBmHsuL6omgnL
pGUWFmEVtg1OFx2a6UrNBAzC2HF9UTUTlknLLCzCKmwbnC46NNOVmgkYhLHj+qJqJiyTlllYhFXYNjhddGimKzUTMAhjx/VF1UxYJi2zsAirsG1wuujQTFdq
JmAQxo7ri6qZsExaZmERVmHb4PlFr4dmulYzAYMwdlxeFMujlknLLCzCKmwbnC46NNO1mgkYhLHj+qJqJiyTlllYhFXYNjhddGimazUTMAhjx/VF1UxYJi2z
sAirsG1wuujQTNdqJmAQxo7ri6qZsExaZmERVmHb4HTRoZmu1UzAIIwd1xdVM2GZtMzCIqzCtsHpokMzXauZgEEYO64vqmbCMmmZhUVYhW2D00WHZrpWMwGD
MHZcX1TNhGXSMguLsArbBqeLDs10rWYCBmHsuL6omgnLpGUWFmEVtg1OFx2a6VrNBAzC2HF9UTUTlknLLCzCKmwbnC46NNO1mgkYhLHj+qJqJiyTlllYhFXY
Nnh+0ZuhmW7UTMAgjB2XF8XyqGXSMguLsArbBqeLDs10o2YCBmHsuL6omgnLpGUWFmEVtg1OFx2a6UbNBAzC2HF9UTUTlknLLCzCKmwbnC46NNONmgkYhLHj
+qJqJiyTlllYhFXYNjhddGimGzUTMAhjx/VF1UxYJi2zsAirsG1wuujQTDdqJmAQxo7ri6qZsExaZmERVmHb4HTRoZlu1EzAIIwd1xdVM2GZtMzCIqzCtsHp
okMz3aiZgEEYO64vqmbCMmmZhUVYhW2D00WHZrpRMwGDMHZcX1TNhGXSMguLsArbBqeLDs10o2YCBmHsuL6omgnLpGUWFmEVtg2eX/R2aKZbNRMwCGPH5UWx
PGqZtMzCIqzCtsHpokMz3aqZgEEYO64vqmbCMmmZhUVYhW2D00WHZrpVMwGDMHZcX1TNhGXSMguLsArbBqeLDs10q2YCBmHsuL6omgnLpGUWFmEVtg1OFx2a
6VbNBAzC2HF9UTUTlknLLCzCKmwbnC46NNOtmgkYhLHj+qJqJiyTlllYhFXYNjhddGimWzUTMAhjx/VF1UxYJi2zsAirsG1wuujQTLdqJmAQxo7ri6qZsExa
ZmERVmHb4HTRoZlu1UzAIIwd1xdVM2GZtMzCIqzCtsHpokMz3aqZgEEYO64vqmbCMmmZhUVYhW2D5xe9G5rpTs0EDMLYcXlRLI9aJi2zsAirsG1wuujQTHdq
JmAQxo7ri6qZsExaZmERVmHb4HTRoZnu1EzAIIwd1xdVM2GZtMzCIqzCtsHpokMz3amZgEEYO64vqmbCMmmZhUVYhW2D00WHZrpTMwGDMHZcX1TNhGXSMguL
sArbBqeLDs10p2YCBmHsuL6omgnLpGUWFmEVtg1OFx2a6U7NBAzC2HF9UTUTlknLLCzCKmwbnC46NNOdmgkYhLHj+qJqJiyTlllYhFXYNjhddGimOzUTMAhj
x/VF1UxYJi2zsAirsG1wuujQTHdqJmAQxo7ri6qZsExaZmERVmHb4PlF74dmulczAYMwdlxeFMujlknLLCzCKmwbnC46NNO9mgkYhLHj+qJqJiyTlllYhFXY
NjhddGimezUTMAhjx/VF1UxYJi2zsAirsG1wuujQTPdqJmAQxo7ri6qZsExaZmERVmHb4HTRoZnu1UzAIIwd1xdVM2GZtMzCIqzCtsHpokMz3auZgEEYO64v
qmbCMmmZhUVYhW2D00WHZrpXMwGDMHZcX1TNhGXSMguLsArbBqeLDs10r2YCBmHsuL6omgnLpGUWFmEVtg1OFx2a6V7NBAzC2HF9UTUTlknLLCzCKmwbnC46
NNO9mgkYhLHj+qJqJiyTlllYhFXYNvjrRb/+8MPj43N8eH745vffvfzr/3p4++bl3988vfvw6tunn9+9DD7+96pzevXhX98/fv/yGR9+l69ef+yHH55+ie+f
fopPv7z7w1ef31De/fTz8z8+fvjw8NfH/sY/vX//9H5848Pbt0+//PHtw7u/vXyGX716/Oj/8ub57Yv+l1fh6cef3j4+P35++8vp3z0/vn/18OrfHt7+/Pjq
L4/Pvzw+vnv1+tXDu+9eXbx+/Q8vX+F//PQy/e7x2zc/Prz96tXTT4/vH54/Tj//4pcv5vun9z/+/Pbh4pvXv/+6//jLWw/fvLyf/vbDy4XOv/bdLf50+N2f
/hNvEX9+/+nD9Ut8ecOrbx/evXt6frnFq3ePf3150789/sOr//nh8eUoL1/Dy6l+fPP28cPz07tHHeev7x8fXi77Lz88vPt/3//pX39+0d2d/s57pMPv0n/i
Pf70/ctn9Hz6vvj009Ut/pO+6OkNH775/U8vn/U/Prz/65uXPzlvH79/+YPz+h8+/p0y79/89Yf+k+ennz59KX95en5++vHTD394fPju8f3HX/Di3z89PX/5
ydcvH+WXp/d/+/Qn9Jv/H1BLAwQUAAAACADOcRhdduwABXoDAABPFAAADQAAAHhsL3N0eWxlcy54bWzdWFtv2jAU/iuR37dAQjMyEaQ2EtKkbarUPuzVEAcs
OZc5poP9+vnYIQmQU4UWNmmJUGwff9+5xnaYVWov2NOGMeXsMpFXEdkoVX523Wq1YRmtPhYly7UkLWRGle7KtVuVktGkAlAmXG80CtyM8pzMZ/k2W2SqclbF
NlcRGTdDjn18SfRgMCGOpYuLhEVkr68PWfYhSYg7n7k1x3yWFnlLFRI7oAlpxpwXKiISU8GXkgMqpRkXezvswcCqEIV0lPaBgR16pPptxWPbA/dqnoznhTS6
rYZTPfeSUwHyZc3QKpDrZURGo/HCn2q3ulomlxI2yNEQ5GWzLzF8/C7Chbkut/CIZGruLkk4hINjRp3zDTPqAsKugeYBBcyFaAr4jtiB+aykSjGZL3THYMzg
mcip28/7UlfwWtL92LsjgwFVIXgCKtdxf77dDvSdpDrhXhxfm9SDGyU1Dx3jZSETJpsoe+QwNJ8JlioNl3y9gacqSijcQqki042E03WRU5OCA6KLdMzKGBG1
MSvbUfofFnAb22BqrWMgwsw15gwE6JkHuwci7OSOY3VDx2vFhHgCkh/p0TK9SztL9AgW6Lxp6kjXTUtjO6Coy2a5u7TBm3idkr8U6mGrXchN/+e2UOxRspTv
TH+XNgZg7OObsns3ZfdxdlqWYn8v+DrPmE3cYIXzGT3gnBcmFV/BqqVLizi/JC2f2U7VC527S3HjJjd1/a5l97rsF/j5hrRdh72TNv/67MFNbf/Usk+uzz79
VzVz9ddlpbtMDnhj6jPuzaI7/lu5wxaj99jejc11+N165+lsb0ebWzPqwCEvIt/h60O0FM5yy4Xied3b8CRh+dkep+kVXeqvpSN+PT9hKd0K9dwII9K2v7GE
b7OwmfUIbtWz2vZXOBToDfNw0NS6eJ6wHUviuqt3+aPzkb0AcCppj+DnEgxjZf0SkGF6MAswjEVhev4nf6aoP1aG2TbtlUxRzBTFjJqPlFNJbG5MTz8m1Fe/
p2Ho+0GARdR+GJxZEGNxCwL49bNhtgEC0wOaLos1nm28Ql6vAyynr1UI5ileiZineKxB0h83QIRhf7YxPYDAsoDVDujv1wM11Y/xfcgqZhv2BuOSMMQkUIv9
NRoESHQCuPvzg70lvh+G/RKQ9Vvg+5gE3kZcglkANmAS3zf74Ml+5B72Kbf9C3H+B1BLAwQUAAAACADOcRhdl4q7HMAAAAATAgAACwAAAF9yZWxzLy5yZWxz
nZK5bsMwDEB/xdCeMAfQIYgzZfEWBPkBVqIP2BIFikWdv6/apXGQCxl5PTwS3B5pQO04pLaLqRj9EFJpWtW4AUi2JY9pzpFCrtQsHjWH0kBE22NDsFosPkAu
GWa3vWQWp3OkV4hc152lPdsvT0FvgK86THFCaUhLMw7wzdJ/MvfzDDVF5UojlVsaeNPl/nbgSdGhIlgWmkXJ06IdpX8dx/aQ0+mvYyK0elvo+XFoVAqO3GMl
jHFitP41gskP7H4AUEsDBBQAAAAIAM5xGF16AH1uaQEAANADAAAPAAAAeGwvd29ya2Jvb2sueG1stZPBbtswDIZfxdADzGmaFVhQ97JibYBhK9qid8Wmay6S
aJB0s/XpR8swZqBAsEtOMn8K1Mef9PWR+LAnOhS/Y0hSuU6135al1B1EL5+oh2SZljh6tZBfS+kZfCMdgMZQrlerqzJ6TO7meq71wOUyIIVakZKJo/CCcJR/
+TEs3lBwjwH1T+XydwBXREwY8R2ayq1cIR0d74nxnZL68FQzhVC5iynxAqxYf5CfRshnv5esqN8/egOp3NXKCrbIovlGru+N8Q3s8hQNSt8wKPCtV7hjGnpM
r2MZ66JctJF9mM/JxC3/j43UtljDLdVDhKSTjwxhBEzSYS+uSD5C5XZJlIfsoIxt2Tu7ZmpRjW1hGG/RErxrMuX5iGykv2ykC5j1CZj1eWEeQWjgGpbeXJ7A
uTwvzrOXwxJlcwJlk5dp3qAGWkzQ/LAyYrptc/3AxXjkca83ny++2NYOIXw17Wf6Tr6ZF3L+mW7+AlBLAwQUAAAACADOcRhdAWXF7sAAAACrAwAAGgAAAHhs
L19yZWxzL3dvcmtib29rLnhtbC5yZWxzxZM5DsIwEEWvYvkADCSBAhEqmrQoF7DMZBHxIs8gkttjoAiWKGhQKuuP5fdfMT6ccVDcO0td70mMZrBUyo7Z7wFI
d2gUrZxHG28aF4ziGEMLXumrahGy9XoH4ZMhj4dPpqgnj78QXdP0Gk9O3wxa/gKGuwtX6hBZilqFFrmUMA7zmOB1bFaRLEV1KWWoLhsJSwtliVC2vFCeCOXL
CxWJUPFHIeJpQJpt3jmp3/6xnuNbnNtf8T1Mt3b3dIDkbx4fUEsDBBQAAAAIAM5xGF2OsKfWJwEAAGcFAAATAAAAW0NvbnRlbnRfVHlwZXNdLnhtbM2Uz07D
MAzGX6XqdWoyBuKA1l2AK+zAC4TWXaPmn2JvdG+P226TQKNiKhK7NGpsfz/Hn5Ll2z4AJq01DvO0JgoPUmJRg1UofADHkcpHq4h/40YGVTRqA3Ixn9/LwjsC
Rxl1Gulq+QSV2hpKnlveRu1dnkYwmCaPQ2LHylMVgtGFIo7LnSu/UbIDQXBln4O1DjjjhFSeJXSRnwGHutcdxKhLSNYq0ouynCVbI5H2BlCMS5zp0VeVLqD0
xdZyicAQQZVYA5A1YhCdjZOJJwzD92Yyv5cZA3LmOvqA7FiEy3FHS7rqLLAQRNLjRzwRWXry+aBzu4Tyl2we74ePTe8Hyn6ZPuOvHp/0L+xjcSV93F5JH3f/
2Me7981fX/1uFVZpd+TL/n1dfQJQSwECFAMUAAAACADNcRhdRsdNSJUAAADNAAAAEAAAAAAAAAAAAAAAgAEAAAAAZG9jUHJvcHMvYXBwLnhtbFBLAQIUAxQA
AAAIAM1xGF38s+k8+QAAACsCAAARAAAAAAAAAAAAAACAAcMAAABkb2NQcm9wcy9jb3JlLnhtbFBLAQIUAxQAAAAIAM1xGF2ZXJwjEAYAAJwnAAATAAAAAAAA
AAAAAACAAesBAAB4bC90aGVtZS90aGVtZTEueG1sUEsBAhQDFAAAAAgAzXEYXYxqPZmLBQAAHxEAABgAAAAAAAAAAAAAAICBLAgAAHhsL3dvcmtzaGVldHMv
c2hlZXQxLnhtbFBLAQIUAxQAAAAIAM1xGF0zKULwCAQAAGAMAAAYAAAAAAAAAAAAAACAge0NAAB4bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWxQSwECFAMUAAAA
CADNcRhdyQtMu+gBAAB7BAAAGAAAAAAAAAAAAAAAgIErEgAAeGwvd29ya3NoZWV0cy9zaGVldDMueG1sUEsBAhQDFAAAAAgAznEYXZo9RF31LAAAOhMCABgA
AAAAAAAAAAAAAICBSRQAAHhsL3dvcmtzaGVldHMvc2hlZXQ0LnhtbFBLAQIUAxQAAAAIAM5xGF127AAFegMAAE8UAAANAAAAAAAAAAAAAACAAXRBAAB4bC9z
dHlsZXMueG1sUEsBAhQDFAAAAAgAznEYXZeKuxzAAAAAEwIAAAsAAAAAAAAAAAAAAIABGUUAAF9yZWxzLy5yZWxzUEsBAhQDFAAAAAgAznEYXXoAfW5pAQAA
0AMAAA8AAAAAAAAAAAAAAIABAkYAAHhsL3dvcmtib29rLnhtbFBLAQIUAxQAAAAIAM5xGF0BZcXuwAAAAKsDAAAaAAAAAAAAAAAAAACAAZhHAAB4bC9fcmVs
cy93b3JrYm9vay54bWwucmVsc1BLAQIUAxQAAAAIAM5xGF2OsKfWJwEAAGcFAAATAAAAAAAAAAAAAACAAZBIAABbQ29udGVudF9UeXBlc10ueG1sUEsFBgAA
AAAMAAwAEAMAAOhJAAAAAA==
"""

def _load_template():
    raw = base64.b64decode("".join(_TEMPLATE_B64.split()))
    return io.BytesIO(raw)

# ---------------------------------------------------------------------------
# Start the JVM with the mpxj jars on the classpath
# ---------------------------------------------------------------------------
import jpype

if not jpype.isJVMStarted():
    try:
        import mpxj
        lib_dir = os.path.join(os.path.dirname(mpxj.__file__), "lib")
        jars = glob.glob(os.path.join(lib_dir, "*.jar"))
    except Exception:
        jars = []
    if not jars:
        raise RuntimeError("Could not locate mpxj jars. Install with: pip install mpxj")
    jpype.startJVM(classpath=jars)

UniversalProjectReader = jpype.JClass("org.mpxj.reader.UniversalProjectReader")

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

HOURS_PER_DAY = 8

# Authoritative active resource roles (bookableresourcecategory names) exactly as
# they exist in Dataverse. The plugin matches role names case-insensitively but
# otherwise exactly, so these strings are what must land in the workbook.
CANONICAL_ROLES = [
    "L1-BP", "L1-CARD PERSO", "L1-IT", "L1-MANAGED SERVICE", "L1-NANO-PERSO",
    "L1-NANO-SWITCH", "L1-PMO", "L1-SOFTWARE DEV", "L1-SUPPORT",
    "L2-BP", "L2-CARD PERSO", "L2-IT", "L2-MANAGED SERVICE", "L2-NANO-PERSO",
    "L2-NANO-SWITCH", "L2-PMO", "L2-Senior QC Manager", "L2-SOFTWARE DEV", "L2-SUPPORT",
    "L3-BP", "L3-CARD PERSO", "L3-IT", "L3-MANAGED SERVICE", "L3-NANO-PERSO",
    "L3-NANO-SWITCH", "L3- PMO", "L3-SOFTWARE DEV", "L3-SUPPORT",
    "L4-BP", "L4-CARD PERSO", "L4-IT", "L4-MANAGED SERVICE", "L4-NANO-PERSO",
    "L4-NANO-SWITCH", "L4-PMO", "L4-SOFTWARE DEV", "L4-SUPPORT",
    "External", "Project Manager",
]

# Discipline spelling variants -> the canonical discipline token used in the role
# names above. Lets long forms ("Software Development") and abbreviations resolve
# to the exact Dataverse spelling ("SOFTWARE DEV").
_DISCIPLINE_ALIASES = {
    "BP": ["BP", "BUSINESS PARTNER", "BUSINESS PROCESS"],
    "CARD PERSO": ["CARD PERSO", "CARD PERSONALIZATION", "CARD PERSONALISATION", "CARDPERSO"],
    "IT": ["IT", "INFORMATION TECHNOLOGY"],
    "MANAGED SERVICE": ["MANAGED SERVICE", "MANAGED SERVICES", "MS"],
    "NANO-PERSO": ["NANO PERSO", "NANO PERSONALIZATION", "NANO PERSONALISATION", "NANOPERSO", "NANO-PERSO"],
    "NANO-SWITCH": ["NANO SWITCH", "NANOSWITCH", "NANO-SWITCH", "NS"],
    "PMO": ["PMO", "PROJECT MANAGEMENT", "PROJECT MANAGEMENT OFFICE"],
    "SOFTWARE DEV": ["SOFTWARE DEV", "SOFTWARE DEVELOPMENT", "SOFTWARE DEVELOPER",
                     "SW DEV", "SWDEV", "DEV", "SD"],
    "SUPPORT": ["SUPPORT", "SUPP"],
    "SENIOR QC MANAGER": ["SENIOR QC MANAGER", "SR QC MANAGER", "SENIOR QC MGR", "QC MANAGER"],
}


def _norm(s):
    """Uppercase; _, -, / -> space; collapse whitespace. Matching only."""
    s = s.upper().replace("_", " ")
    s = re.sub(r"[-/]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


_ALIAS_TO_DISCIPLINE = {}
for _disc, _aliases in _DISCIPLINE_ALIASES.items():
    for _a in _aliases:
        _ALIAS_TO_DISCIPLINE[_norm(_a)] = _disc

# (level, canonical-discipline) -> exact Dataverse role string.
_ROLE_BY_KEY = {}
for _role in CANONICAL_ROLES:
    _m = re.match(r"^L\s*([1-4])\s*-?\s*(.*)$", _role, re.I)
    if not _m:
        continue
    _dn = _norm(_m.group(2))
    _disc = _ALIAS_TO_DISCIPLINE.get(_dn, _dn)
    _ROLE_BY_KEY[(_m.group(1), _disc)] = _role


def _canonical_role(name):
    """
    Resolve an L<1-4> label to its exact Dataverse role string, else None.
    Handles any spacing/case, "-"/"/" separators, and discipline spelling
    variants (e.g. "L3 - Software Development" -> "L3-SOFTWARE DEV").
    """
    if name is None:
        return None
    m = re.match(r"^\s*L\s*([1-4])\b[\s\-/]*(.*)$", str(name).strip(), re.I)
    if not m:
        return None
    rest = _norm(m.group(2))
    if not rest:
        return None
    disc = _ALIAS_TO_DISCIPLINE.get(rest, rest)
    return _ROLE_BY_KEY.get((m.group(1), disc))


def format_resource_name(resource_name):
    """
    Canonical label for BOTH the Resources sheet 'Resource Name' column and the
    Tasks sheet 'Resources' column, so they always agree. An L<1-4> role becomes
    its exact Dataverse spelling; anything else is returned unchanged.
    """
    if resource_name is None or str(resource_name).strip() == "":
        return resource_name
    canon = _canonical_role(resource_name)
    return canon if canon is not None else str(resource_name).strip()


def map_role(resource_name):
    """
    Value for the Resources sheet 'Role' column. An L<1-4> role maps to its exact
    Dataverse role string; anything else -> 'External' (which is a real role).
    """
    if resource_name is None or str(resource_name).strip() == "":
        return None
    canon = _canonical_role(resource_name)
    return canon if canon is not None else "External"


def read_project(mpp_path):
    return UniversalProjectReader().read(mpp_path)


def work_hours(task):
    w = task.getWork()
    if w is None:
        return 0.0
    try:
        TimeUnit = jpype.JClass("org.mpxj.TimeUnit")
        return float(w.convertUnits(TimeUnit.HOURS, None).getDuration())
    except Exception:
        try:
            return float(w.getDuration())
        except Exception:
            return 0.0


def duration_days(task):
    d = task.getDuration()
    if d is None:
        return None
    try:
        TimeUnit = jpype.JClass("org.mpxj.TimeUnit")
        return float(d.convertUnits(TimeUnit.DAYS, None).getDuration())
    except Exception:
        try:
            return float(d.getDuration())
        except Exception:
            return None


def is_milestone(task):
    try:
        return bool(task.getMilestone())
    except Exception:
        return False


def resource_names(task):
    out = []
    for a in task.getResourceAssignments():
        r = a.getResource()
        if r is not None and r.getName() is not None:
            n = format_resource_name(str(r.getName()).strip())
            if n and n not in out:
                out.append(n)
    return out


def percent_complete(task):
    try:
        p = task.getPercentageComplete()
        if p is None:
            return None
        return int(round(float(p)))
    except Exception:
        return None


def fmt_date(dt):
    """
    Convert an mpxj date/datetime to a Python datetime, preserving time-of-day
    so Start/Finish match what MS Project shows (e.g. 8:00 AM / 5:00 PM).
    Returns None when absent. Written as a real Excel datetime and formatted
    MPP-style at write time.
    """
    if dt is None:
        return None
    s = str(dt).strip()
    if not s:
        return None
    # Normalise a trailing "Z" and space/T separator, then let fromisoformat try.
    iso = s.replace("Z", "").replace(" ", "T")
    try:
        return datetime.datetime.fromisoformat(iso)
    except ValueError:
        pass
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(iso[:19], fmt)
        except ValueError:
            continue
    try:
        return datetime.datetime.strptime(s[:10], "%Y-%m-%d")
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Flatten the task tree
# ---------------------------------------------------------------------------
def build_rows(project):
    tasks = [t for t in project.getTasks()
             if t is not None and t.getName() is not None]

    def has_children(t):
        c = t.getChildTasks()
        return c is not None and not c.isEmpty()

    rows = []
    id_map = {}
    resource_order = []
    counter = [1]
    pending_preds = {}

    def add_resource(n):
        if n not in resource_order:
            resource_order.append(n)

    def emit(t, parent_template_id):
        this_id = counter[0]
        counter[0] += 1
        id_map[int(t.getID())] = this_id

        summary = has_children(t)
        row = {
            "Task ID": this_id,
            "Parent Task ID": parent_template_id,
            "Task Name": str(t.getName())[:300],
            "Start": None, "Duration (days)": None, "Finish": None,
            "Effort (hours)": None, "% Complete": None,
            "Predecessors": None, "Resources": None, "Notes": None,
        }

        if not summary:
            row["Start"] = fmt_date(t.getStart())
            row["Finish"] = fmt_date(t.getFinish())
            if is_milestone(t):
                row["Duration (days)"] = 0
            else:
                dd = duration_days(t)
                if dd is None:
                    wh = work_hours(t)
                    dd = round(wh / HOURS_PER_DAY, 2) if wh else None
                row["Duration (days)"] = round(dd, 2) if dd is not None else None
            wh = work_hours(t)
            row["Effort (hours)"] = round(wh, 2) if wh else None
            row["% Complete"] = percent_complete(t)

            rns = resource_names(t)
            for n in rns:
                add_resource(n)
            row["Resources"] = "; ".join(rns) if rns else None

            preds = []
            try:
                for rel in t.getPredecessors():
                    src = rel.getPredecessorTask()
                    if src is not None:
                        preds.append(int(src.getID()))
            except Exception:
                pass
            if preds:
                pending_preds[this_id] = preds

        rows.append(row)
        if summary:
            for c in t.getChildTasks():
                if c is not None and c.getName() is not None:
                    emit(c, this_id)
        return this_id

    min_level = min(int(t.getOutlineLevel()) for t in tasks)
    roots = [t for t in tasks if int(t.getOutlineLevel()) == min_level]
    while len(roots) == 1 and has_children(roots[0]):
        roots = [c for c in roots[0].getChildTasks()
                 if c is not None and c.getName() is not None]
    for r in roots:
        emit(r, None)

    row_by_id = {r["Task ID"]: r for r in rows}
    for tid, raw_preds in pending_preds.items():
        mapped = [str(id_map[p]) for p in raw_preds if p in id_map]
        if mapped:
            row_by_id[tid]["Predecessors"] = "; ".join(mapped)

    return rows, resource_order


def project_estimated_start(project, rows):
    """
    Estimated start date for the project as a whole (project-level, not a
    per-task or per-resource value). Prefers the project's own Start Date
    property (the "Start date" shown in MS Project's Project Information
    dialog); falls back to the earliest task start date in the schedule if
    that property isn't set on the file.
    """
    try:
        props = project.getProjectProperties()
        dt = fmt_date(props.getStartDate())
        if dt is not None:
            return dt
    except Exception:
        pass
    starts = [r["Start"] for r in rows if r["Start"] is not None]
    return min(starts) if starts else None


# ---------------------------------------------------------------------------
# Fill the embedded template (keeps inline strings -> plugin can read it)
# ---------------------------------------------------------------------------
def write_excel(rows, resources, out_path, project_name="Imported Project",
                 project_start=None):
    wb = openpyxl.load_workbook(_load_template())

    # --- Project: set the Project Name value (row with Setting == Project Name)
    ws_p = wb["Project"]
    for r in range(2, ws_p.max_row + 1):
        if str(ws_p.cell(r, 1).value).strip().lower() == "project name":
            ws_p.cell(r, 2, project_name)
            break

    # --- Project: set the Estimated Start Date (project-level setting, added
    # below the template's existing rows; not on the Resources or Tasks sheet)
    estimated_start_row = None
    for r in range(2, ws_p.max_row + 1):
        if str(ws_p.cell(r, 1).value).strip().lower() == "estimated start date":
            estimated_start_row = r
            break
    if estimated_start_row is None:
        estimated_start_row = ws_p.max_row + 1
        ws_p.cell(estimated_start_row, 1, "Estimated Start Date")
        ws_p.cell(estimated_start_row, 3,
                  "Start date of the project, taken from the MPP file.")
    if project_start is not None:
        ws_p.cell(estimated_start_row, 2, project_start)
        ws_p.cell(estimated_start_row, 2).number_format = "ddd m/d/yy h:mm AM/PM"

    # --- Resources
    ws_r = wb["Resources"]
    # clear anything below header just in case
    for r in range(2, ws_r.max_row + 1):
        for c in range(1, 4):
            ws_r.cell(r, c, None)
    for i, name in enumerate(resources, start=2):
        ws_r.cell(i, 1, name)
        ws_r.cell(i, 2, None)            # blank -> generic team member
        ws_r.cell(i, 3, map_role(name))  # preserved role mapping

    # --- Tasks
    ws_t = wb["Tasks"]
    headers = ["Task ID", "Parent Task ID", "Task Name", "Start",
               "Duration (days)", "Finish", "Effort (hours)", "% Complete",
               "Predecessors", "Resources", "Notes"]
    # clear old rows
    for r in range(2, ws_t.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws_t.cell(r, c, None)

    r = 2
    for row in rows:
        for c, h in enumerate(headers, start=1):
            ws_t.cell(r, c, row[h])
        # MPP-style display: e.g. "Mon 6/1/26 8:00 AM". Underlying value is a
        # real Excel datetime, so it stays sortable/computable.
        _mpp_fmt = "ddd m/d/yy h:mm AM/PM"
        if row["Start"] is not None:
            ws_t.cell(r, 4).number_format = _mpp_fmt
        if row["Finish"] is not None:
            ws_t.cell(r, 6).number_format = _mpp_fmt
        r += 1
    last_row = max(r - 1, 2)

    # Remove any pre-existing table, then add the Estimates table over the data
    try:
        if "Estimates" in ws_t.tables:
            del ws_t.tables["Estimates"]
    except Exception:
        pass
    table = Table(displayName="Estimates", ref="A1:K{}".format(last_row))
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws_t.add_table(table)

    wb.save(out_path)


# ---------------------------------------------------------------------------
def main():
    try:
        from google.colab import files  # type: ignore
        in_colab = True
    except Exception:
        in_colab = False

    if in_colab:
        print("Please upload your .mpp file...")
        uploaded = files.upload()
        if not uploaded:
            print("No file uploaded. Aborting.")
            return
        mpp_path = list(uploaded.keys())[0]
        base = os.path.splitext(os.path.basename(mpp_path))[0]
        out_path = base + "_Import.xlsx"

        project = read_project(mpp_path)
        rows, resources = build_rows(project)
        project_start = project_estimated_start(project, rows)
        write_excel(rows, resources, out_path, project_name=base, project_start=project_start)
        print("Wrote {} task rows, {} resources -> {}".format(len(rows), len(resources), out_path))
        files.download(out_path)
        return

    if len(sys.argv) < 2:
        print("Usage: python mpp_to_import.py <input.mpp> [output.xlsx]")
        sys.exit(1)
    mpp_path = sys.argv[1]
    base = os.path.splitext(os.path.basename(mpp_path))[0]
    out_path = sys.argv[2] if len(sys.argv) >= 3 else base + "_Import.xlsx"
    project = read_project(mpp_path)
    rows, resources = build_rows(project)
    project_start = project_estimated_start(project, rows)
    write_excel(rows, resources, out_path, project_name=base, project_start=project_start)
    print("Wrote {} task rows, {} resources -> {}".format(len(rows), len(resources), out_path))


if __name__ == "__main__":
    main()